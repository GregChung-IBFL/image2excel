"""
Processing module for image2excel.  Handles one image file end-to-end.
To use, instantiate a Converter, then call its process_file method.  A
Converter is for single-use only, use a new Converter to process another image.

Coded by Greg Chung : https://github.com/GregChung-IBFL/image2excel
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from PIL import Image   # pillow package from https://python-pillow.org/


class Converter :
    """The main worker class for the image-to-spreadsheet conversion."""

    class SheetInfo :
        """Mini data class representing the size of the spreadsheet image."""
        def __init__(self) :
            self.num_sheet_cols = self.num_sheet_rows = 1

    def __init__( self, filepath:str, config:dict ) :
        self.filepath = filepath
        self.config = config
        self.sheet_info = Converter.SheetInfo()
        # image_info will contain tuples of (label, text) string describing the image and
        # image file, which will be used to populate the Information worksheet in the spreadsheet.
        self.image_info = []
        self._add_image_info( "File & Image Properties:" )

    def _add_image_info( self, label:str = "", value:str = "" ) :
        """Helper function to append (@label, @value) string tuples into the image_info list."""
        self.image_info.append( (label, value) )


    def calc_resize_dimensions( self, image:Image, output_width:int, output_height:int, allow_enlarge:bool ) :
        """Calculates and returns the dimensions of the resized-to-fit @image.  (Does not
        modify @image.)
        Returns tuple containing the (new width, new height, factor) which best fits the desired
        @output_width and @output_height.  The new dimensions maintain the aspect ratio
        of the original @image.  Images will only be enlarged to fit if @allow_enlarge is True.
        """
        # Calculate the rescaling factor for each dimension to reach the desired output size.
        # If the original image is smaller than the desired size, the factor will exceed 100%.
        horz_rescale = output_width / image.width
        vert_rescale = output_height / image.height

        # However, if enlarging of images is not allowed, cap the scaling factors at 1.0.
        if not allow_enlarge :
            horz_rescale = min( 1.0, horz_rescale )
            vert_rescale = min( 1.0, vert_rescale )

        # Resizing maintains aspect ratio in both dimensions, so the overall scaling
        # factor is the minimum of either the X or Y scaling.
        rescale = min( horz_rescale, vert_rescale )
        return ( int(round(image.width * rescale)), int(round(image.height * rescale)), rescale )


    def convert_to_RGB_mode( self, image ) :
        """Converts the image to RGB mode.  Mainly useful for GIF paletted format."""
        if image.mode != "RGB" :
            image = image.convert( mode = "RGB", dither = None )
        return image


    def resize_image_to_fit( self, image:Image, new_dimensions ) -> Image:
        """Resizes @image to fit desired output @new_dimensions.
        Returns the resized image.
        """
        image = image.resize( new_dimensions[:2] ) # Omit the rescale factor at the end

        self._add_image_info( "Resizing Factor", "{:.1%}".format(new_dimensions[2]) )
        self._add_image_info( "Resized Resolution (W x H)", "%d x %d" % new_dimensions[:2] )

        return image


    def intialize_image_sheet( self, wkbook ) :
        """Applies various styles to the Image worksheet."""
        wksheet = wkbook.active
        wksheet.title = "Image"

        # The global settings are more efficient than trying to set these per row!  Especially,
        # hiding rows using openpyxl's group() function seems to apply the settings individually.
        wksheet.sheet_format.defaultColWidth = self.config["output_col_width"]
        wksheet.sheet_format.defaultRowHeight = self.config["output_row_height"]
        wksheet.sheet_format.customHeight = True
        wksheet.sheet_format.zeroHeight = 1
        wksheet.sheet_view.showGridLines = False
        wksheet.sheet_view.showRowColHeaders = False

        # Hide the columns beyond the image.
        wksheet.column_dimensions.group( start = get_column_letter(self.sheet_info.num_sheet_cols + 1), end = get_column_letter(16384), hidden = True )

        # The initial zoom of the sheet, by default = 100%.  For very large images, this might
        # need to be set lower.  For small images, this could be higher.
        wksheet.sheet_view.zoomScale = self.config["output_zoom"]

        return wksheet


    def populate_cell_values( self, wksheet, image ) :
        """Populates the @wksheet with the color values making up @image.
        Loop through every pixel in @image and write the appropriate numeric values
        (0-255) into the corresponding spreadsheet cells:
            1. Get the RGB values from getpixel(), which returns a (r, g, b) tuple
            2. Set the values into the corresponding cells
        """

        # Version 1 uses getpixel().  getdata might be faster.

        for row in range( self.sheet_info.num_sheet_rows ) :
            # Reminder, addresses in Excel are 1-indexed while the image and other data is
            # 0-indexed.  Thus +1 is added where setting addresses.
            wksheet.row_dimensions[ row + 1 ].height = self.config["output_row_height"]

            pixel_rgb = (0,0,0)
            for col in range( self.sheet_info.num_sheet_cols ) :
                pixel_x = col // 3    # Three sheet cols per pixel col
                rgb_index = col % 3   # R,G,B
                if rgb_index == 0 :
                    pixel_coord = ( pixel_x, row )
                    pixel_rgb = image.getpixel( pixel_coord )  # Returns (R,G,B) tuple

                cell = wksheet.cell( column = col + 1, row = row + 1 )
                cell.value = pixel_rgb[ rgb_index ]


    def apply_color_rules( self, wksheet ) :
        """Applies the conditional formatting rules to the @wksheet which give color to the cells.
        Every spreadsheet column is assigned a 2-Color Scale style Graded Color Scale conditional
        format, where the gradient ranges from black (cell value = 0) to full red, green, or blue
        (cell value = 255).  The rules are created, and then assigned to the alternating columns.
        """

        # First, define the rules representing the three colors.
        rules = []
        csr_attrs = { "start_type" : "num", "start_value" : 0, "end_type" : "num", "end_value" : 255, "start_color" : "00000000" }
        for end_color in [ "00FF0000", "0000FF00", "000000FF" ] :
            rules.append( ColorScaleRule( **csr_attrs, end_color = end_color ) )

        # Second, generate the conditional formatting ranges.  Ranges use absolute addressing,
        # e.g. A1:A150.  OpenPyXL does not seem to allow conditional formatting using a whole
        # column notation, A:A.
        cell_ranges = [ [], [], [] ]
        for col in range( self.sheet_info.num_sheet_cols ) :
            letter = get_column_letter(col + 1)
            rgb_index = col % 3
            cell_ranges[rgb_index].append( F"{letter}1:{letter}{self.sheet_info.num_sheet_rows}" )

        # Third, apply the ranges to the rules.
        for index in range ( 3 ) :
            wksheet.conditional_formatting.add( " ".join(cell_ranges[index]), rules[index] )


    def write_info_sheet( self, wkbook ) :
        """Creates and populates the Information worksheet into @wkbook.
        Writes the contents of the image_info list, plus the various program settings from config.
        Remember that config will contain the default settings from the configuration file, plus
        any values applied from the command line arguments.
        """
        infosheet = wkbook.create_sheet( "Information" )
        infosheet.column_dimensions[ "A" ].width = infosheet.column_dimensions[ "B" ].width = 35

        # Merge in the program settings from config, with the exception of the presets definitions.
        self._add_image_info()
        self._add_image_info( "Program Arguments & Configuration:" )
        for key, value in self.config.items() :
            if key == "presets" :
                continue
            self._add_image_info( key, str(value) )

        # Dump the combined file properties + settings tuples into the sheet.
        row = 3
        for label, value in self.image_info :
            infosheet.cell( column = 1, row = row, value = label )
            infosheet.cell( column = 2, row = row, value = value )
            row += 1



    def get_output_file_path( self, input_file_path, output_file_path ) -> str:
        """Returns the file path for the output file.
        If @output_file_name is specified, its value will be used as the file name for the
        output spreadsheet file.  If not specified, the output file name will be the same
        as the image file, but with the ".xlsx" extension.
        By default, the file will be created under the same directory as the image file.
        """
        if ( output_file_path is None or output_file_path == "" ) :
            # splitext returns the tuple, e.g. ("C:\Text\Example\MyFile", ".jpg"),
            # so simply swap the file extension.
            output_file_path = os.path.splitext( input_file_path )[0] + ".xlsx"
        return output_file_path



    def save_excel_file( self, wkbook, output_file_path:str ) :
        """Saves the spreadsheet represented by @wkbook to the file named by @output_file_path.
        The output file must be writeable.  Any exceptions along the way are reported and eaten.
        """
        try :
            print( F'Saving spreadsheet to "{output_file_path}"...' )
            wkbook.save( output_file_path )
        except FileNotFoundError :
            print( "FileNotFoundError: Please verify path is valid." )
        except PermissionError :
            print( "PermissionError: File may exist and be read-only, or is already open elsewhere." )
        except Exception as exc:
            print( "Something bad happened saving the file: " + str(exc) )



    def process_file(self) :
        """Converts one image file to a spreadsheet representation.
        """

        # PIL blocks extremely large files, but by default the limit is around 89.5 megapixels.
        Image.MAX_IMAGE_PIXELS = 6000 * 4000  # 24 MP

        try :
            print( F'Loading image file "{self.filepath}"...' )
            # file must remain open for as long as PIL may need to read from it. In this
            # code, that includes up through the resizing step, after which image will be
            # the resized copy (in memory).
            with open( self.filepath, "br" ) as file :
                image = Image.open( file )

                file_stats = os.stat( self.filepath )
                self._add_image_info( "Image File", self.filepath)
                self._add_image_info( "File Size", "{:,}".format(file_stats.st_size))
                self._add_image_info( "Format", image.format)
                self._add_image_info( "Original Resolution (W x H)", "%d x %d" % image.size)

                # I am not sanity checking for "reasonable" image sizes.  An image one pixel
                # tall but 15000 pixels wide might cause problems?  Out of scope for this demo.

                new_dimensions = self.calc_resize_dimensions( image,
                                                  output_width = self.config["output_width"],
                                                  output_height = self.config["output_height"],
                                                  allow_enlarge = self.config.get("enlarge") )

                image = self.convert_to_RGB_mode( image )

                # Resize the image to fit the desired dimensions.
                print( "Resizing image (%d x %d) to (%d x %d)..." % (image.size + new_dimensions[:2]) )
                image = self.resize_image_to_fit( image, new_dimensions )

        except FileNotFoundError :
            print( "FileNotFoundError:  Please verify path is valid and file exists." )
            return
        except Image.DecompressionBombError :
            print( "DecompressionBombError:  Image is too large to process, please try a smaller image." )
            return

        # Remember the dimensions of the spreadsheet "image" area:
        self.sheet_info.num_sheet_cols = image.width * 3
        self.sheet_info.num_sheet_rows = image.height

        # Figure out the outfile file path & name.
        output_file_path = self.get_output_file_path( self.filepath, self.config["output_file"] )

        self._add_image_info( "Spreadsheet Range", "A1:{lastcol}{lastrow}".format( lastcol = get_column_letter(self.sheet_info.num_sheet_cols), lastrow = self.sheet_info.num_sheet_rows ) )
        self._add_image_info( "Spreadsheet File", output_file_path )
        self._add_image_info( "Generated at", datetime.now().strftime( "%b %d %Y %I:%M:%S %p") )

        # The real work commenceth!  Start by initializing an empty workbook.
        wkbook = Workbook()

        # Initialize the first worksheet, which will contain the picture.
        wksheet = self.intialize_image_sheet( wkbook )

        # Populate the "Image" worksheet.
        self.populate_cell_values( wksheet, image )

        # Apply the conditional formatting which displays the colors.
        self.apply_color_rules( wksheet )

        # Create and populate the Information worksheet with details about the image.
        self.write_info_sheet( wkbook )

        # Write the output file, then we're done.
        self.save_excel_file( wkbook, output_file_path )
